import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date
import os

# ── Colors ────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")  # auto filled
YELLOW = PatternFill("solid", fgColor="FFEB9C")  # needs engineer
RED    = PatternFill("solid", fgColor="FFC7CE")  # mismatch/error


def safe_write(ws, row, col, value, fill=None):
    """
    Safely writes to a cell.
    Handles merged cells by only writing to top-left cell.
    """
    try:
        cell = ws.cell(row=row, column=col)
        
        # Check if this is a slave merged cell
        # by trying to get its coordinate and checking
        # against all merged ranges
        cell_coord = cell.coordinate
        is_slave = False
        
        for merged_range in ws.merged_cells.ranges:
            range_str = str(merged_range)
            
            # Get the top-left cell of this merged range
            top_left = ws.cell(
                row=merged_range.min_row, 
                column=merged_range.min_col
            ).coordinate
            
            # If our cell is in this range but not the top-left
            if (cell_coord != top_left and 
                cell.row >= merged_range.min_row and
                cell.row <= merged_range.max_row and
                cell.column >= merged_range.min_col and
                cell.column <= merged_range.max_col):
                is_slave = True
                break
        
        if not is_slave:
            cell.value = value
            if fill:
                cell.fill = fill
                
    except Exception as e:
        # If anything goes wrong, just skip this cell
        print(f"  ⚠️  Skipped cell row={row} col={col}: {e}")


def write_bom(extracted_data, template_path, output_path):
    """
    Main function — fills the BOM template with extracted data.
    
    extracted_data format:
    {
        "customer": "MSIL",
        "rfq_no": "RFQ-2026-001",
        "date": "2026-04-10",
        "parts": [
            {
                "serial_no": 1,
                "part_number": "72450-69Q00",
                "part_name": "Silencer Comp Dash Outer",
                "drawing_no": "72450-69QX0_DEC 19 2025",
                "cad_data_no": "M569887-72450-69Q00",
                "project_model": "YVF-NM",
                "production_volume": 60173,
                "production_location": "SMG GUJARAT",
                "raw_material": "EVAC",
                "gsm": 1100,
                "color_code": "Not Available",
                "folded_L": 1018,
                "folded_W": 457,
                "folded_H": 185,
                "blank_L": 1053,
                "blank_W": 578,
                "blank_T": 20,
                "surface_area_sqm": 0.344065,
                "cad_volume_mm3": None,
                "weight_cad_g": 0,
                "weight_drawing_g": 380,
                "weight_bom_g": 355,
                "mfg_process": None,
                "tool_layout": None,
                "assy_process": None,
                "qty_assy": None,
                "remarks": "Calculated GSM is 1100 APPROX"
            }
        ]
    }
    """
    
    print(f"📂 Loading template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    
    # Fill each sheet
    _fill_part_summary(wb, extracted_data)
    _fill_bom_sheet(wb, extracted_data)
    
    # Save output
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ BOM saved to: {output_path}")
    return output_path


def _fill_part_summary(wb, data):
    """
    Fills the Part Summary sheet.
    
    Template structure:
    Row 3: Customer in F3
    Row 4: Date in J4  
    Row 5-6: Headers (merged, don't touch)
    Row 7+: Data rows — one per part
    
    Columns:
    B=Serial No, C=Part Number, D=Part Name, 
    E=Drawing No, F=CAD Data No, G=Project/Model,
    H=Production Volume, I=Location, J=RFQ No
    """
    
    ws = wb["Part Summary"]
    parts = data.get("parts", [])
    
    # Write customer name — cell F3
    safe_write(ws, row=3, col=6, 
               value=data.get("customer", ""), 
               fill=GREEN)
    
    # Write date — cell J4
    safe_write(ws, row=4, col=10, 
               value=data.get("date", str(date.today())), 
               fill=GREEN)
    
    # Write each part — starting at row 7
    for i, part in enumerate(parts):
        row = 7 + i
        
        safe_write(ws, row, col=2,  
                   value=part.get("serial_no", i+1),        
                   fill=GREEN)
        safe_write(ws, row, col=3,  
                   value=part.get("part_number"),            
                   fill=GREEN)
        safe_write(ws, row, col=4,  
                   value=part.get("part_name"),              
                   fill=GREEN)
        safe_write(ws, row, col=5,  
                   value=part.get("drawing_no"),             
                   fill=GREEN)
        safe_write(ws, row, col=6,  
                   value=part.get("cad_data_no"),            
                   fill=GREEN)
        safe_write(ws, row, col=7,  
                   value=part.get("project_model"),          
                   fill=GREEN)
        safe_write(ws, row, col=8,  
                   value=part.get("production_volume"),      
                   fill=GREEN)
        safe_write(ws, row, col=9,  
                   value=part.get("production_location"),    
                   fill=GREEN)
        safe_write(ws, row, col=10, 
                   value=data.get("rfq_no"),                 
                   fill=GREEN)
        
        print(f"  ✅ Part Summary row {row}: "
              f"{part.get('part_number')} written")


def _fill_bom_sheet(wb, data):
    """
    Fills the BOM sheet.
    
    Template structure (from our analysis):
    Rows 2-6:  Part headers (pull from Part Summary via formula)
    Row 7:     Date/RFQ row
    Row 8:     Column headers
    Row 9:     Sub-headers (L/W/H etc)
    Row 10+:   Data rows — one per part
    
    Columns (all offset from B=col2):
    B=Item#, C=Part Name, D=Raw Material, E=GSM,
    F=Folded L, G=Folded W, H=Folded H,
    I=Blank L, J=Blank W, K=Blank T,
    L=Color Code, M=Texturing, N=CAD Volume,
    O=Surface Area, P=Weight CAD, Q=Weight Drawing,
    R=Weight BOM, S=MFG Process, T=Tool Layout,
    U=Assy Process, V=Qty/Assy, W=Remarks
    """
    
    ws = wb["BOM"]
    parts = data.get("parts", [])
    
    # Write each part — starting at row 10
    for i, part in enumerate(parts):
        row = 10 + i
        
        # ── Auto-filled fields (GREEN) ─────────────────────
        
        safe_write(ws, row, col=2,  # Item #
                   value=part.get("serial_no", i+1),
                   fill=GREEN)
        
        safe_write(ws, row, col=4,  # Raw Material
                   value=part.get("raw_material", "TBC"),
                   fill=GREEN if part.get("raw_material") else YELLOW)
        
        safe_write(ws, row, col=5,  # GSM
                   value=part.get("gsm"),
                   fill=GREEN if part.get("gsm") else YELLOW)
        
        # Folded dimensions
        safe_write(ws, row, col=6,  # Folded L
                   value=part.get("folded_L"),
                   fill=GREEN if part.get("folded_L") else YELLOW)
        safe_write(ws, row, col=7,  # Folded W
                   value=part.get("folded_W"),
                   fill=GREEN if part.get("folded_W") else YELLOW)
        safe_write(ws, row, col=8,  # Folded H
                   value=part.get("folded_H"),
                   fill=GREEN if part.get("folded_H") else YELLOW)
        
        # Blank/Unfolded dimensions
        safe_write(ws, row, col=9,  # Blank L
                   value=part.get("blank_L"),
                   fill=GREEN if part.get("blank_L") else YELLOW)
        safe_write(ws, row, col=10, # Blank W
                   value=part.get("blank_W"),
                   fill=GREEN if part.get("blank_W") else YELLOW)
        safe_write(ws, row, col=11, # Blank T
                   value=part.get("blank_T"),
                   fill=GREEN if part.get("blank_T") else YELLOW)
        
        safe_write(ws, row, col=12, # Color Code
                   value=part.get("color_code", "TBC"),
                   fill=GREEN if part.get("color_code") else YELLOW)
        
        safe_write(ws, row, col=13, # Texturing
                   value="Not Applicable",
                   fill=GREEN)
        
        safe_write(ws, row, col=14, # CAD Volume mm3
                   value=part.get("cad_volume_mm3"),
                   fill=GREEN if part.get("cad_volume_mm3") else YELLOW)
        
        safe_write(ws, row, col=15, # Surface Area SQM
                   value=part.get("surface_area_sqm"),
                   fill=GREEN if part.get("surface_area_sqm") else YELLOW)
        
        safe_write(ws, row, col=16, # Weight CAD
                   value=part.get("weight_cad_g", 0),
                   fill=GREEN)
        
        safe_write(ws, row, col=17, # Weight Drawing
                   value=part.get("weight_drawing_g"),
                   fill=GREEN if part.get("weight_drawing_g") else YELLOW)
        
        safe_write(ws, row, col=18, # Weight BOM
                   value=part.get("weight_bom_g"),
                   fill=GREEN if part.get("weight_bom_g") else YELLOW)
        
        safe_write(ws, row, col=23, # Remarks
                   value=part.get("remarks", ""),
                   fill=GREEN)
        
        # ── Engineer fills these (YELLOW) ─────────────────
        
        safe_write(ws, row, col=19, # MFG Process
                   value=part.get("mfg_process", "TBC"),
                   fill=GREEN if part.get("mfg_process") else YELLOW)
        
        safe_write(ws, row, col=20, # Tool Layout
                   value=part.get("tool_layout", "TBC"),
                   fill=YELLOW)
        
        safe_write(ws, row, col=21, # Assy Process
                   value=part.get("assy_process", "TBC"),
                   fill=YELLOW)
        
        safe_write(ws, row, col=22, # Qty/Assy
                   value=part.get("qty_assy", "TBC"),
                   fill=YELLOW)
        
        # ── GSM Cross Check ────────────────────────────────
        _gsm_crosscheck(ws, row, part)
        
        print(f"  ✅ BOM row {row}: "
              f"{part.get('part_number')} written")


def _gsm_crosscheck(ws, row, part):
    """
    Validates GSM from drawing against 
    calculated GSM from weight and surface area.
    Flags red if more than 10% difference.
    """
    
    gsm          = part.get("gsm")
    surface_area = part.get("surface_area_sqm")
    weight       = part.get("weight_drawing_g")
    
    if not all([gsm, surface_area, weight]):
        return
    
    calculated_gsm = weight / surface_area
    diff_pct = abs(calculated_gsm - gsm) / gsm * 100
    
    if diff_pct > 10:
        # Flag remarks cell red with warning
        remarks_cell = ws.cell(row=row, column=23)
        existing = remarks_cell.value or ""
        remarks_cell.value = (
            f"{existing} | ⚠️ GSM mismatch: "
            f"drawing={gsm}, "
            f"calculated={round(calculated_gsm)}"
        )
        remarks_cell.fill = RED
        print(f"  ⚠️  GSM mismatch flagged at row {row}")
    else:
        print(f"  ✅ GSM check passed: "
              f"drawing={gsm}, "
              f"calculated={round(calculated_gsm)}")


# ── Test with your real template data ─────────────────────
if __name__ == "__main__":
    
    # This matches the real data in your MSIL template
    test_data = {
        "customer": "MSIL",
        "rfq_no": "RFQ-2026-61564-P003",
        "date": "2026-04-10",
        "parts": [
            {
                "serial_no": 1,
                "part_number": "72450-69Q00",
                "part_name": "Silencer Comp Dash Outer",
                "drawing_no": "72450-69QX0_DEC 19 2025",
                "cad_data_no": "M569887-72450-69Q00-RRFQ_000.001",
                "project_model": "YVF-NM",
                "production_volume": 60173,
                "production_location": "SMG GUJARAT",
                "raw_material": "EVAC",
                "gsm": 1100,
                "color_code": "Not Available",
                "folded_L": 1018,
                "folded_W": 457,
                "folded_H": 185,
                "blank_L": 1053,
                "blank_W": 578,
                "blank_T": 20,
                "surface_area_sqm": 0.344065,
                "cad_volume_mm3": None,
                "weight_cad_g": 0,
                "weight_drawing_g": 380,
                "weight_bom_g": 355,
                "mfg_process": None,
                "tool_layout": None,
                "assy_process": None,
                "qty_assy": None,
                "remarks": "Calculated GSM is 1100 APPROX"
            },
            {
                "serial_no": 2,
                "part_number": "72450-69Q10",
                "part_name": "Silencer Comp Dah Outer - CNG",
                "drawing_no": "72450-69QX1_DEC 22 2025",
                "cad_data_no": "M569887-72450-69Q10-RRFQ_000.001",
                "project_model": "YVF-NM",
                "production_volume": 39010,
                "production_location": "SMG GUJARAT",
                "raw_material": "EVAC",
                "gsm": 1100,
                "color_code": "Not Available",
                "folded_L": 1018,
                "folded_W": 458,
                "folded_H": 185,
                "blank_L": 1053,
                "blank_W": 578,
                "blank_T": 20,
                "surface_area_sqm": 0.341674,
                "cad_volume_mm3": None,
                "weight_cad_g": 0,
                "weight_drawing_g": 380,
                "weight_bom_g": 375.9,
                "mfg_process": None,
                "tool_layout": None,
                "assy_process": None,
                "qty_assy": None,
                "remarks": "Calculated GSM is 1100 APPROX"
            },
            {
                "serial_no": 3,
                "part_number": "72450-69QA0",
                "part_name": "Silencer Comp Dah Outer - SHEV",
                "drawing_no": "72450-69QX2_DEC 24 2025",
                "cad_data_no": "M569887-72450-69QA0-RRFQ_000.001",
                "project_model": "YVF-NM",
                "production_volume": 12000,
                "production_location": "SMG GUJARAT",
                "raw_material": "EVAC",
                "gsm": 1550,
                "color_code": "Not Available",
                "folded_L": 1018,
                "folded_W": 457,
                "folded_H": 184,
                "blank_L": 1053,
                "blank_W": 578,
                "blank_T": 20,
                "surface_area_sqm": 0.355261,
                "cad_volume_mm3": None,
                "weight_cad_g": 0,
                "weight_drawing_g": 380,
                "weight_bom_g": 540,
                "mfg_process": None,
                "tool_layout": None,
                "assy_process": None,
                "qty_assy": None,
                "remarks": "Calculated GSM is 1550 APPROX"
            }
        ]
    }
    
    print("🚀 Starting BOM writer test...")
    print(f"   Parts to write: {len(test_data['parts'])}")
    print()
    
    write_bom(
        extracted_data = test_data,
        template_path  = "templates/bom_template.xlsx",
        output_path    = "output/test_bom_output.xlsx"
    )
    
    print()
    print("📊 Open output/test_bom_output.xlsx to check the result")
    print("   GREEN  = auto filled by tool")
    print("   YELLOW = engineer needs to fill")
    print("   RED    = GSM mismatch — check this!")
