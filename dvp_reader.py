"""
dvp_reader.py
─────────────────────────────────────────────────────────────
Reads an RFQ folder and generates a DVP Test Plan.
"""

import anthropic
import base64
import json
import os
import re
from PIL import Image
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

Image.MAX_IMAGE_PIXELS = None
client = anthropic.Anthropic()

STANDARDS_LIBRARY_PATH = "standards_library"


# ── Standards Library ──────────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    text = re.sub(r'\(\d+\)', '', text)
    text = re.sub(r'[^a-zA-Z0-9]', '', text)
    return text.lower()


def build_library_index(library_path: str) -> dict:
    index = {}
    if not os.path.exists(library_path):
        print(f"  ⚠️  Standards library not found: {library_path}")
        os.makedirs(library_path, exist_ok=True)
        return index
    for filename in os.listdir(library_path):
        if filename.startswith("."):
            continue
        base = os.path.splitext(filename)[0]
        key  = _normalize(base)
        index[key] = filename
    print(f"  📚 Standards library: {len(index)} files indexed")
    return index


def check_availability(standard: str, library_index: dict) -> bool:
    norm_std = _normalize(standard)
    base_std = re.sub(r'[a-z0-9]{1,2}$', '', norm_std)
    for key in library_index:
        if norm_std == key:
            return True
        if norm_std in key and len(norm_std) >= 6:
            return True
        if key in norm_std and len(key) >= 6:
            return True
        if base_std and (base_std in key or key in base_std):
            return True
        # Match customer-provided specs like tsl3608g14p1 → TSL3608G
        key_base = re.sub(r'\d+p\d+$', '', key)  # strip version suffix
        if norm_std in key_base or key_base in norm_std:
            return True
    return False


# ── Claude helpers ─────────────────────────────────────────────────────────────

def _img_to_b64(img: Image.Image, quality: int = 85) -> str:
    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="JPEG", quality=quality)
    return base64.standard_b64encode(buf.getvalue()).decode()


def _ask_vision(b64: str, prompt: str) -> str:
    msg = client.messages.create(
        model      = "claude-sonnet-4-20250514",
        max_tokens = 2000,
        messages   = [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                {"type": "text", "text": prompt}
            ]
        }]
    )
    return msg.content[0].text.strip()


def _ask_text(prompt: str) -> str:
    msg = client.messages.create(
        model      = "claude-sonnet-4-20250514",
        max_tokens = 2000,
        messages   = [{"role": "user", "content": prompt}]
    )
    return msg.content[0].text.strip()


def _parse_json(raw: str) -> dict:
    raw = re.sub(r'^```json\s*', '', raw.strip())
    raw = re.sub(r'\s*```$', '', raw)
    if not raw:
        return {}
    return json.loads(raw)


def _clean_code(code: str) -> str:
    """Fix common OCR misreads."""
    code = re.sub(r'\bT5L', 'TSL', code)
    code = re.sub(r'\bTS(\d)', r'TSL\1', code)
    code = re.sub(r'\bV\s(\d)', r'VW \1', code)  # V 01155 → VW 01155
    return code.strip()


def _merge_variants(all_standards: dict) -> dict:
    """
    Merge variant entries into one.
    e.g. TSL3608G-1B, TSL3608G-2B, TSL3608G-3B, TSL3608G-4C → TSL3608G-1B, 2B, 3B, 4C
    """
    base_groups = {}
    for code, s in all_standards.items():
        # Normalize "or" variants: "1B, 2B, 3B or 4C" → "1B, 2B, 3B, 4C"
        code = re.sub(r'\s+or\s+', ', ', code, flags=re.IGNORECASE)
        base = re.sub(r'[-/]\w+$', '', code).strip()
        if base not in base_groups:
            base_groups[base] = {"codes": [], "s": s}
        base_groups[base]["codes"].append(code)

    final_standards = {}
    for base, group in base_groups.items():
        codes = group["codes"]
        if len(codes) == 1:
            final_standards[codes[0]] = group["s"]
        else:
            variants = []
            for c in codes:
                suffix = re.sub(r'.*[-/]', '', c).strip()
                if suffix and suffix.lower() != base.lower():
                    variants.append(suffix)

            if variants:
                merged_code = f"{base}-{', '.join(variants)}"
            else:
                merged_code = base

            merged = group["s"].copy()
            merged["code"] = merged_code
            final_standards[merged_code] = merged

    return final_standards


def _extract_criteria_from_spec(pdf_path: str, standard_code: str) -> str:
    """
    Read spec PDF and extract key criteria in one line.
    Only reads first 4 pages — summary/properties table is always early.
    """
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:8]:
                text += (page.extract_text() or "")

        if not text.strip():
            return f"As per {standard_code}"

        prompt = f"""
You are reading a test standard document: {standard_code}

Extract KEY test criteria in ONE concise line from the properties/requirements table.
Include important numerical values and limits.

If this is an environmental/substances control document with no numerical test criteria,
return: "Comply with {standard_code} environmental/substance requirements"

If this is a MATERIAL standard with a properties/requirements table (Table 2 or similar),
extract the key numerical limits from that table for the most relevant class/grade.
Focus on: density, surface density, bending strength, tensile strength, flammability, sound absorption.
Example: "Density 0.1±0.01 g/cm³, Surface density 1000±100 g/m², Bending strength ≥4N/50mm, Tensile ≥150kPa, Flammability per TSM0504G Method A, Sound absorption ≥Fig.6"

Examples:
- "Density 0.1±0.01 g/cm³, Bending strength ≥ 4N/50mm, Flammability per TSM0504G"
- "Insertion loss per FIG.1, Sample 1000×1200mm, N=3, Freq 400Hz-10kHz"
- "Sound absorption coefficient ≥ Fig.6 values, Sample φ99mm & φ29mm, Freq 200-10000Hz"

If this is a test METHOD standard (defines how to test, no pass/fail limits),
summarize ALL tests covered and key conditions in one line.
Example: "Insertion loss & sound absorption coefficient per drawing Fig.1, Sample 1000×1200mm, N=3, Freq 400Hz-10kHz"
Do NOT include thickness variation, mass variation, soaking time, or temperature/humidity conditions.

Return ONLY the one-line criteria string, nothing else.
Do NOT explain what the document is. Do NOT say "Looking at this document".
Do NOT include any preamble. Return ONLY the criteria value itself.
If no numerical criteria exist, return: "Comply with [standard_code] requirements"

SPEC TEXT:
{text[:12000]}
"""
        criteria = _ask_text(prompt).strip()
        criteria = criteria.strip('"\'').replace('\n', ' ')
        return criteria

    except Exception as e:
        return f"As per {standard_code}"


# ── Google Vision OCR ──────────────────────────────────────────────────────────

def _google_ocr(img: Image.Image) -> str:
    from google.cloud import vision
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "rfq_vision_key.json"
    client_vision = vision.ImageAnnotatorClient()

    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="JPEG", quality=95)
    content = buf.getvalue()

    image    = vision.Image(content=content)
    response = client_vision.text_detection(image=image)

    if response.error.message:
        print(f"   ⚠️  Google Vision error: {response.error.message}")
        return ""

    texts = response.text_annotations
    if texts:
        return texts[0].description
    return ""


# ── PDF Classifier ─────────────────────────────────────────────────────────────

def _classify_pdfs(folder_path: str, all_pdfs: list) -> tuple:
    import pdfplumber

    STANDARD_PATTERN = re.compile(
        r'\b(TSL|ISO|DIN|VW|TL|PV|DBL|GMW|FORD|FLTM|SAE|ASTM|JIS|BS)\s*[\d]',
        re.IGNORECASE
    )
    CAD_KEYWORDS = [
        "projection", "datum", "tolerance", "scale", "drawn by",
        "third angle", "first angle", "quotation drawing", "rev",
        "section", "detail", "assembly", "zeichnung"
    ]
    SKIP_NAME_HINTS = ["cpl", "rfq", "quotation", "commercial", "price"]
    SKIP_CONTENT_HINTS = [
        "injection molding", "statement of requirement",
        "purging", "raw material", "molding process",
        "barrel", "masterbatch"
    ]
    SPEC_NAME_HINTS = ["tdo", "tcp", "tl_", "spec", "test", "requirement"]
    TOYOTA_STD_PATTERN = re.compile(r'^(TSL|TSM|TSZ|TLL|TSS)', re.IGNORECASE)

    drawing_pdfs = []
    spec_files   = []

    for f in all_pdfs:
        full_path = os.path.join(folder_path, f)
        size      = os.path.getsize(full_path)
        fl        = f.lower()

        if any(k in fl for k in SKIP_NAME_HINTS):
            print(f"   ⏭️  Skipping (irrelevant doc): {f}")
            continue

        print(f"   🔍 Classifying: {f} ({size//1000}KB)...")

        try:
            text       = ""
            text_len   = 0
            text_lower = ""
            with pdfplumber.open(full_path) as pdf:
                for page in pdf.pages[:2]:
                    text += (page.extract_text() or "")

            text_len   = len(text.strip())
            text_lower = text.lower()

            # Skip process/manufacturing documents
            if text_len > 100 and any(k in text_lower for k in SKIP_CONTENT_HINTS):
                print(f"   ⏭️  Skipping (process doc, not test standard): {f}")
                continue

            if TOYOTA_STD_PATTERN.match(f):
                print(f"   📋 Spec doc (Toyota standard filename): {f}")
                spec_files.append(f)
                continue

            if text_len < 150:
                if any(k in fl for k in SPEC_NAME_HINTS):
                    print(f"   📋 Spec doc (name hint + image-based): {f}")
                    spec_files.append(f)
                elif size > 500_000:
                    print(f"   📐 Drawing (image-based, large file): {f}")
                    drawing_pdfs.append(f)
                else:
                    print(f"   ⏭️  Skipping (image-based, small file): {f}")
                continue

            if any(k in text_lower for k in CAD_KEYWORDS):
                print(f"   📐 Drawing (CAD keywords found): {f}")
                drawing_pdfs.append(f)
                continue

            if STANDARD_PATTERN.search(text):
                print(f"   📋 Spec doc (test standards found): {f}")
                spec_files.append(f)
                continue

            if size > 300_000:
                print(f"   📐 Drawing (large file, no text signals): {f}")
                drawing_pdfs.append(f)
            else:
                print(f"   ⏭️  Skipping (no relevant content): {f}")

        except Exception as e:
            print(f"   ⚠️  Could not read {f}: {e}")

    return drawing_pdfs, spec_files


# ── Drawing Prompt ─────────────────────────────────────────────────────────────

DRAWING_PROMPT = """
You are reading an automotive engineering drawing.
Extract ALL test standards AND their specific acceptance criteria
mentioned anywhere in the drawing.

IMPORTANT: All output must be in ENGLISH only.
Translate any German, Japanese, or other language text into English.
For example:
- "Werkstoffkennwerte" → "Material properties"
- "Brandsicherheit nach Flammwidrigkeit" → "Fire safety / Flammability"
- "Zeichnungsangaben nach Schnittstelle" → "Drawing specifications per interface"
- "Recyclingbeleg" → "Recycling compliance"

For each standard, look nearby for the specific value/criteria.
e.g. "DIN EN ISO 845 — 30 ± 3 kg/m³" → criteria: "30 ± 3 kg/m³"
     "TL 1010" with no value nearby    → criteria: ""
     "PV 1200 — 6 cycles"             → criteria: "6 cycles"
     "TSL0605G"                        → criteria: ""

Also check NOTES sections carefully. Standards may appear mid-sentence with their requirement described around them.
Extract the full requirement as criteria.
e.g. Note: "TO COMPLY WITH NONFLAMMABILITY CRITERION (SES N 3245). SHALL NOT USE HBCD & DBDE."
→ code: "SES N 3245", criteria: "Nonflammability criterion, shall not use HBCD & DBDE"

e.g. Note: "CONFORM TO THE RESTRICTIONS ON USE OF ENVIRONMENTAL IMPACT SUBSTANCES SPECIFIED IN SES N 2402."
→ code: "SES N 2402", criteria: "Restrictions on use of environmental impact substances"

e.g. Note: "TO COMPLY WITH VOLATILE ORGANIC COMPOUNDS EVAPORATIVE CRITERION (SES N 2403)."
→ code: "SES N 2403", criteria: "Volatile organic compounds evaporative criterion"

e.g. Note: "TO COMPLY WITH ODOR SENSORY EVALUATION IN SES N 2405."
→ code: "SES N 2405", criteria: "Odor sensory evaluation"

Return ONLY valid JSON — no markdown:
{
    "part_number": "",
    "part_name": "",
    "customer": "",
    "spec_reference": "",
    "standards": [
        {
            "code": "DIN EN ISO 845",
            "description": "Gross density test",
            "criteria": "30 ± 3 kg/m³",
            "context": "material specification table"
            "description" must be the specific test name (e.g. "Gross density test", "Flammability test"), 
            NOT the section header (e.g. NOT "specification list", NOT "material specifications").
            If you cannot determine the test name from the standard code, use the standard's known purpose.

        }
    ]
}
"""


# ── Step 1: Extract from Drawing TIF ──────────────────────────────────────────

def extract_from_drawing(tif_path: str) -> dict:
    """Extract standards from drawing TIF using Claude Vision."""
    print(f"\n📐 Reading drawing: {os.path.basename(tif_path)}")

    img = Image.open(tif_path)
    W, H = img.size
    print(f"   Size: {W}×{H}px")

    img = img.crop((0, int(H * 0.20), W, H))
    W, H = img.size

    cols = 4 if W > 15000 else 3
    rows = 3
    row_overlap = int(H * 0.08)
    col_overlap = int(W * 0.08)
    all_standards = {}
    result = {
        "part_number":    "",
        "part_name":      "",
        "customer":       "",
        "spec_reference": "",
        "standards":      [],
    }

    for row in range(rows):
        for col in range(cols):
            y1 = max(0, int(H * row / rows) - row_overlap)
            y2 = min(H, int(H * (row+1) / rows) + row_overlap)
            x1 = max(0, int(W * col / cols) - col_overlap)
            x2 = min(W, int(W * (col+1) / cols) + col_overlap)
            crop = img.crop((x1, y1, x2, y2))
            crop.thumbnail((4000, 4000), Image.LANCZOS)

            print(f"   🔍 Cell [{row+1},{col+1}]...")

            try:
                ocr_text = _google_ocr(crop)
                if not ocr_text.strip():
                    print(f"   ⏭️  Cell [{row+1},{col+1}] empty OCR — skipping")
                    continue
                prompt = f"{DRAWING_PROMPT}\n\nDRAWING TEXT:\n{ocr_text}"
                raw    = _ask_text(prompt)
                data   = _parse_json(raw)

                if not result["part_number"] and data.get("part_number"):
                    result["part_number"] = data["part_number"]
                if not result["part_name"] and data.get("part_name"):
                    result["part_name"] = data["part_name"]
                if not result["customer"] and data.get("customer"):
                    result["customer"] = data["customer"]
                if not result["spec_reference"] and data.get("spec_reference"):
                    result["spec_reference"] = data["spec_reference"]

                for s in (data.get("standards") or []):
                    code = s.get("code", "").strip()
                    if not code:
                        continue
                    if re.match(r'^\d[A-Z0-9]{2}[\.\s]\d{3}[\.\s]\d{3}', code):
                        continue
                    if re.match(r'^\d+$', code):
                        continue
                    if not re.search(r'\d', code):
                        continue
                    if code not in all_standards:
                        all_standards[code] = s

            except Exception as e:
                print(f"   ⚠️  Cell [{row+1},{col+1}] error: {e}")

    # ── Clean OCR misreads — runs ONCE after all strips
    cleaned = {}
    for code, s in all_standards.items():
        clean = _clean_code(code)
        s["code"] = clean
        if clean not in cleaned:
            cleaned[clean] = s
    all_standards = cleaned

    # ── Merge variants — runs ONCE after all strips
    all_standards = _merge_variants(all_standards)

    result["standards"] = list(all_standards.values())
    print(f"   ✅ Found {len(result['standards'])} standards in drawing")
    print(f"      {[s['code'] for s in result['standards']]}")
    return result


# ── Step 1b: Extract from Drawing PDF ─────────────────────────────────────────

def extract_from_drawing_pdf(pdf_path: str) -> dict:
    """
    Extract standards from a PDF drawing (Toyota, Honda etc.)
    Uses Google Vision OCR + Claude for understanding.
    Scans full page in 3x3 grid — works for any customer layout.
    """
    print(f"\n📐 Reading PDF drawing: {os.path.basename(pdf_path)}")
    print(f"   Size: checking pages...")

    result = {
        "part_number":    "",
        "part_name":      "",
        "customer":       "",
        "spec_reference": "",
        "standards":      [],
    }
    all_standards = {}

    try:
        from pdf2image import convert_from_path

        pages = convert_from_path(pdf_path, dpi=300)
        print(f"   📃 {len(pages)} pages found")

        for page_num, page_img in enumerate(pages):
            cW, cH = page_img.size
            print(f"   🖼  Page {page_num+1}: {cW}×{cH}px")

            rows, cols  = 3, 3
            row_overlap = int(cH * 0.08)
            col_overlap = int(cW * 0.08)

            for row in range(rows):
                for col in range(cols):
                    y1 = max(0, int(cH * row / rows) - row_overlap)
                    y2 = min(cH, int(cH * (row+1) / rows) + row_overlap)
                    x1 = max(0, int(cW * col / cols) - col_overlap)
                    x2 = min(cW, int(cW * (col+1) / cols) + col_overlap)

                    cell = page_img.crop((x1, y1, x2, y2))
                    cell.thumbnail((4000, 4000), Image.LANCZOS)

                    print(f"   🔍 Page {page_num+1} Cell [{row+1},{col+1}]...")

                    try:
                        ocr_text = _google_ocr(cell)
                        if not ocr_text.strip():
                            print(f"   ⏭️  Cell [{row+1},{col+1}] empty OCR — skipping")
                            continue
                        prompt   = f"{DRAWING_PROMPT}\n\nDRAWING TEXT:\n{ocr_text}"
                        raw      = _ask_text(prompt)
                        data     = _parse_json(raw)

                        if not result["part_number"] and data.get("part_number"):
                            result["part_number"] = data["part_number"]
                        if not result["part_name"] and data.get("part_name"):
                            result["part_name"] = data["part_name"]
                        if not result["customer"] and data.get("customer"):
                            result["customer"] = data["customer"]
                        if not result["spec_reference"] and data.get("spec_reference"):
                            result["spec_reference"] = data["spec_reference"]

                        for s in (data.get("standards") or []):
                            code = s.get("code", "").strip()
                            if not code:
                                continue
                            if re.match(r'^\d[A-Z0-9]{2}[\.\s]\d{3}[\.\s]\d{3}', code):
                                continue
                            if re.match(r'^\d+$', code):
                                continue
                            if not re.search(r'\d', code):
                                print(f"   🚫 Rejected (no digits): '{code}'")
                                continue
                            if code not in all_standards:
                                all_standards[code] = s

                    except Exception as e:
                        print(f"   ⚠️  Cell [{row+1},{col+1}] error: {e}")

        # ── Clean OCR misreads — runs ONCE after all cells
        cleaned = {}
        for code, s in all_standards.items():
            clean = _clean_code(code)
            s["code"] = clean
            if clean not in cleaned:
                cleaned[clean] = s
        all_standards = cleaned

        # ── Merge variants — runs ONCE after all cells
        all_standards = _merge_variants(all_standards)

        result["standards"] = list(all_standards.values())
        print(f"   ✅ Found {len(result['standards'])} standards")
        print(f"      {[s['code'] for s in result['standards']]}")
        return result

    except ImportError:
        print("   ❌ pdf2image not installed. Run: pip install pdf2image")
        return result
    except Exception as e:
        print(f"   ❌ PDF drawing error: {e}")
        return result


# ── Step 2: Spec Prompt ────────────────────────────────────────────────────────

SPEC_PROMPT = """
You are reading an automotive test specification document.
This could be a VW TDO, Toyota spec, or any customer test document.

Extract the COMPLETE list of tests. For each test, check:
Does this document itself contain HOW TO PERFORM the test?

Return ONLY valid JSON — no markdown:
{
    "spec_number": "",
    "tests": [
        {
            "description": "Flammability test",
            "standard": "TL 1010",
            "responsibility": "Contractor",
            "procedure_in_spec": false,
            "criteria": "test criteria if mentioned"
        }
    ],
    "all_standards_mentioned": ["TL 1010", "VW 50180"] 
}

procedure_in_spec = true  — document contains actual test procedure/conditions.
procedure_in_spec = false — only references an external standard code.
Responsibility: use exactly "Contractor" or "Customer".
"""


def extract_from_spec(pdf_path: str) -> dict:
    print(f"\n📄 Reading spec doc: {os.path.basename(pdf_path)}")
    try:
        import pdfplumber
        full_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            print(f"   📃 Reading all {total} pages...")
            for i in range(total):
                text = pdf.pages[i].extract_text() or ""
                full_text += f"\n--- PAGE {i+1} ---\n{text}"

        if not full_text.strip():
            print(f"   ⚠️  No text found in spec doc — skipping")
            return {"spec_number": "", "tests": [], "all_standards_mentioned": []}

        prompt = f"{SPEC_PROMPT}\n\nSPEC TEXT:\n{full_text}"
        raw    = _ask_text(prompt)

        if not raw.strip():
            print(f"   ⚠️  No response from Claude — skipping spec doc")
            return {"spec_number": "", "tests": [], "all_standards_mentioned": []}

        try:
            result = _parse_json(raw)
        except Exception:
            print(f"   ⚠️  Could not parse Claude response — skipping")
            return {"spec_number": "", "tests": [], "all_standards_mentioned": []}

        if not result:
            return {"spec_number": "", "tests": [], "all_standards_mentioned": []}

        print(f"   ✅ Found {len(result.get('tests', []))} tests in spec")
        return result

    except Exception as e:
        print(f"   ❌ Spec doc error: {e}")
        return {"spec_number": "", "tests": [], "all_standards_mentioned": []}


# ── Step 3: Build DVP test list ───────────────────────────────────────────────

def build_dvp_list(drawing_data:  dict,
                   spec_data:     dict,
                   library_index: dict,
                   folder_path:   str = "",
                   company_name:  str = "") -> list:
    print(f"\n🔧 Building DVP test list...")

    customer = drawing_data.get("customer", "unknown")
    db_key   = re.sub(r'[^a-zA-Z0-9]', '_', customer).lower().strip('_')
    db_path  = f"customer_profiles/criteria_db_{db_key}.json"

    criteria_db = {}
    if os.path.exists(db_path):
        with open(db_path, encoding="utf-8") as f:
            criteria_db = json.load(f)
        print(f"   📋 Loaded criteria for '{customer}': {len(criteria_db)} entries")
    else:
        print(f"   ⚠️  No criteria db for '{customer}' — run build_criteria_db.py first")

    dvp      = []
    seen     = set()
    spec_ref = (drawing_data.get("spec_reference") or spec_data.get("spec_number") or "")

    def _add(serial, description, standard, responsibility, criteria, procedure_in_spec=False):
        key = standard.strip()
        if key in seen:
            return
        seen.add(key)

        if spec_ref and ("Annex" in standard or procedure_in_spec):
            method = f"{standard} (Spec {spec_ref})"
        else:
            method = standard

        if procedure_in_spec:
            available = True
        else:
            available = check_availability(standard, library_index)

        if not criteria:
            criteria = (criteria_db.get(standard.strip()) or
                        criteria_db.get(re.sub(r'\s+', '', standard)) or "")

        print(f"   🔎 {standard.strip()} | available={available} | criteria='{criteria}'")

        # If still no criteria and standard is available, read from spec PDF
        if available:
            # Strip variant suffixes before normalizing e.g. "TSL3608G-1B, 2B, 3B, 4C" → "TSL3608G"
            # Strip variant suffixes only (e.g. "TSL3608G-1B, 2B" → "TSL3608G")
            base_code = re.split(r'[-,]', standard.strip())[0].strip()
            norm_std = _normalize(base_code)
            print(f"      norm_std={norm_std}")
            for lib_key, filename in library_index.items():
                if norm_std in lib_key or lib_key in norm_std:
                    print(f"      matched: {lib_key} → {filename}")
                    for search_path in [STANDARDS_LIBRARY_PATH, folder_path]:
                        full_path = os.path.join(search_path, filename)
                        if os.path.exists(full_path):
                            print(f"   📖 Reading criteria: {filename}")
                            criteria = _extract_criteria_from_spec(full_path, standard.strip())
                            break
                    break

        dvp.append({
            "serial_no":      str(serial),
            "description":    description,
            "method":         method,
            "responsibility": responsibility,
            "criteria":       criteria,
            "available":      available,
        })

    serial = 1

    for test in spec_data.get("tests", []):
        std               = test.get("standard",          "").strip()
        desc              = test.get("description",       "").strip()
        resp              = test.get("responsibility",    "")
        procedure_in_spec = test.get("procedure_in_spec", False)
        criteria          = test.get("criteria",          "")

        if resp.lower() == "contractor":
            resp = company_name.upper() if company_name else "CONTRACTOR"
        elif resp.lower() == "customer":
            resp = customer.upper()
        else:
            resp = ""

        if std and desc:
            _add(serial, desc, std, resp, criteria, procedure_in_spec)
            serial += 1

    for s in drawing_data.get("standards", []):
        code = s.get("code", "").strip()
        if not code or code in seen:
            continue
        if "SPEC" in code.upper() or "TDO" in code.upper():
            continue

        context = s.get("context", "")
        vague_phrases = [
            "referenced", "listed", "specifications table",
            "technical specifications", "material specifications",
            "visible in", "mentioned"
        ]
        if any(p in context.lower() for p in vague_phrases):
            desc = f"Requirements per {code}"
        else:
            desc = (s.get("description", "") or context or f"Requirements per {code}")

        drawing_criteria = s.get("criteria", "")
        _add(serial, desc, code, "", drawing_criteria, False)
        serial += 1

    available_count   = sum(1 for t in dvp if t["available"])
    unavailable_count = len(dvp) - available_count
    print(f"   ✅ Total tests:    {len(dvp)}")
    print(f"   🟢 Available:      {available_count}")
    print(f"   🔴 Not available:  {unavailable_count}")
    return dvp


# ── Step 4: Write DVP Excel ────────────────────────────────────────────────────

def write_dvp_excel(dvp_tests: list, drawing_data: dict, output_path: str, company_name: str = ""):
    print(f"\n📊 Writing Excel: {output_path}")
    print(f"   🏢 company_name in write_dvp_excel: '{company_name}'")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DVP Test Plan"

    RED_BG    = "FF0000"
    GREEN_BG  = "00B050"
    HEADER_BG = "4472C4"

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(row, col, value="", bold=False, size=9, color="000000",
              bg=None, align="left", wrap=True, border_on=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        if border_on:
            c.border = border
        return c

    ws.merge_cells("A1:J1")
    _cell(1, 1, company_name, bold=True, size=12, align="left", border_on=False)
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 8

    ws.merge_cells("A3:C3")
    _cell(3, 1, company_name, bold=True, size=10, border_on=False)
    ws.row_dimensions[3].height = 20

    customer = drawing_data.get("customer", "")
    ws.merge_cells("A4:J4")
    _cell(4, 1, f"DVP Test Plan - {customer}", bold=True, size=12, align="center")
    ws.row_dimensions[4].height = 22

    ws.merge_cells("A5:J5")
    part_info = (
        f"Part: {drawing_data.get('part_name', '')}  |  "
        f"Drawing: {drawing_data.get('part_number', '')}  |  "
        f"Customer: {customer}"
    )
    _cell(5, 1, part_info, bold=True, size=9)
    ws.row_dimensions[5].height = 18

    headers    = ["Serial\nNo.", "Test Description", "Test Method",
                  "Responsibility", "Test Criteria",
                  "Test Start\nDate", "Test End\nDate",
                  "Test\nAgency", "Current", "Remarks"]
    col_widths = [8, 30, 22, 14, 42, 12, 12, 12, 10, 14]

    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _cell(6, i, h, bold=True, size=9, color="FFFFFF", bg=HEADER_BG, align="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[6].height = 30

    for row_num, test in enumerate(dvp_tests, start=7):
        available    = test.get("available", False)
        remarks_text = "AVAILABLE" if available else "NOT AVAILABLE"
        remarks_bg   = GREEN_BG   if available else RED_BG

        row_data = [
            test.get("serial_no",      ""),
            test.get("description",    ""),
            test.get("method",         ""),
            test.get("responsibility", ""),
            test.get("criteria",       ""),
            "", "", "", "",
            remarks_text,
        ]

        for col_i, value in enumerate(row_data, start=1):
            if col_i == 10:
                _cell(row_num, col_i, value, bold=True, size=9,
                      color="FFFFFF", bg=remarks_bg, align="center")
            else:
                _cell(row_num, col_i, value, size=9,
                      align="center" if col_i == 1 else "left")

        ws.row_dimensions[row_num].height = 30

    ws.freeze_panes = "A7"
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    available   = sum(1 for t in dvp_tests if t["available"])
    unavailable = len(dvp_tests) - available
    print(f"   ✅ Saved: {output_path}")
    print(f"   🟢 Available:     {available}")
    print(f"   🔴 Not available: {unavailable}")


# ── Main ───────────────────────────────────────────────────────────────────────

def generate_dvp(folder_path:       str,
                 standards_library: str = STANDARDS_LIBRARY_PATH,
                 output_path:       str = None,
                 company_name:      str = ""):
    print("\n" + "=" * 60)
    print("🔬 DVP PLAN GENERATOR")
    print("=" * 60)

    all_files = [f for f in os.listdir(folder_path) if not f.startswith(".")]

    tif_files = sorted([f for f in all_files if f.lower().endswith((".tif", ".tiff"))])
    all_pdfs  = sorted([f for f in all_files if f.lower().endswith(".pdf")])

    print(f"\n🔎 Classifying {len(all_pdfs)} PDFs (text-only, no Vision)...")
    drawing_pdfs, spec_files = _classify_pdfs(folder_path, all_pdfs)

    drawing_files = tif_files or drawing_pdfs

    print(f"\n   Drawing:   {drawing_files}")
    print(f"   Spec docs: {spec_files}")
    print(f"   Standards: {standards_library}")

    if not drawing_files:
        print("❌ No drawing found (TIF or PDF)!")
        return None

    library_index = build_library_index(standards_library)

    # ── Add customer spec docs to library index
    if spec_files:
        print(f"\n📚 Adding customer spec docs to library index...")
        for spec_file in spec_files:
            base = os.path.splitext(spec_file)[0]
            key  = _normalize(base)
            library_index[key] = spec_file
            print(f"   ✅ {spec_file} → '{key}'")

    # ── Read drawing
    drawing_file = drawing_files[0]
    drawing_path = os.path.join(folder_path, drawing_file)

    if drawing_file.lower().endswith((".tif", ".tiff")):
        drawing_data = extract_from_drawing(drawing_path)
    else:
        drawing_data = extract_from_drawing_pdf(drawing_path)

    spec_data = {"spec_number": "", "tests": [], "all_standards_mentioned": []}

    dvp_tests = build_dvp_list(drawing_data, spec_data, library_index, folder_path, company_name=company_name)

    if not output_path:
        part_no = (drawing_data.get("part_number", "UNKNOWN").replace(".", "_"))
        output_path = os.path.join("output", f"DVP_{part_no}.xlsx")

    write_dvp_excel(dvp_tests, drawing_data, output_path, company_name=company_name)

    json_path = output_path.replace(".xlsx", ".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"drawing": drawing_data, "spec": spec_data, "tests": dvp_tests},
                  f, indent=2, ensure_ascii=False)

    available   = sum(1 for t in dvp_tests if t["available"])
    unavailable = len(dvp_tests) - available

    print("\n" + "=" * 60)
    print("📊 SUMMARY")
    print("=" * 60)
    print(f"Part:        {drawing_data.get('part_name')}")
    print(f"Drawing No:  {drawing_data.get('part_number')}")
    print(f"Customer:    {drawing_data.get('customer')}")
    print(f"Spec Ref:    {drawing_data.get('spec_reference')}")
    print(f"Total tests: {len(dvp_tests)}")
    print(f"🟢 Available:    {available}")
    print(f"🔴 Not avail:    {unavailable}")
    print(f"\nOutput: {output_path}")

    if unavailable > 0:
        print(f"\n⚠️  Standards to request from customer:")
        for t in dvp_tests:
            if not t["available"]:
                print(f"   • {t['method']}")

    return dvp_tests


# ── Run ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    folder = sys.argv[1] if len(sys.argv) > 1 else "rfq_inputs/RFQ_VW_001"
    generate_dvp(
        folder_path       = folder,
        standards_library = "standards_library",
    )